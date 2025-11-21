from sqlalchemy import create_engine
import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import math
import datetime
import glob
from datetime import datetime, timedelta
pd.options.mode.chained_assignment = None
import sys

Diccionario={
#    "ServerDB":"192.168.52.173",
#    "NameDB":"RPA_Netapplications",
#    "UserDB":"NetApplications_RPA",
#    "PasswordDB":"pfU#6$53g*R8",
#    "Schema":'ARIA',
    "ServerDB":"192.168.50.57",
    "NameDB":"RPA",
    "UserDB":"Fabrica_RPA",
    "PasswordDB":"FABRICARPA2020!",
    "Schema":"NotasCreditoYFacturacion",
    "Tabla":"CONSOLIDADO",
    "Hoja":'CIERRE',
    "Ruta":r"\\192.168.50.169\RPA_Notas credito_Facturación\Insumo\Notas Credito\Octubre_2024\02-Oct-2024\Notas antes de cierre_octubre.xlsx",
    "rutaAdjunto":r'\\192.168.50.169\RPA_Notas credito_Facturación\Insumo\Notas Credito\Octubre_2024\02-Oct-2024\Notas antes de cierre_octubre.xlsx',
    "Anexo5OneDrive":r"C:\Users\CGGRPARPM\OneDrive - colsubsidio.com (1)\Anexo 5 - ASCII Negociadores.xlsx",
    "Anexo5":r"\\192.168.50.169\RPA_Notas credito_Facturación\Insumo\Anexos\Anexo 5 - ASCII Negociadores.xlsx",
    "Query":"SELECT * FROM ARIA.CONSOLIDADO",
    "filename":r"\\192.168.50.169\RPA_Notas credito_Facturación\Resultados\Facturas\Facturas_04_2023\CARGUE1",
    "rutaCSV":r"\\192.168.50.169\RPA_Notas credito_Facturación\Resultados\Ejecucion_Marzo_2024\15-Mar-2024\Soportes\Facturas\ListaPedidos.csv",
    "rutaExcel":r"\\192.168.50.169\RPA_Notas credito_Facturación\Resultados\Ejecucion_Marzo_2024\15-Mar-2024\Resultados\ListaPedidos.xlsx",
    "FechaYM":"2024"
}    


# Función para crear una conexión a una base de datos de SQL Server
def ConexionDB(Diccionario):
    try:
        # Establecer la conexión con SQL Server
        server = str(Diccionario.get('ServerDB'))
        database = str(Diccionario.get('NameDB'))
        username = str(Diccionario.get('UserDB'))
        password = str(Diccionario.get('PasswordDB'))
        schema = str(Diccionario.get('Schema')) 
        #database='RPA'
        driver = 'ODBC Driver 17 for SQL Server' # El controlador que estés utilizando
        # Crear la cadena de conexión
        connection_string = f"mssql+pyodbc://{username}:{password}@{server}/{database}?driver={driver}&schema={schema}"
        #connection_string = f"mssql+pyodbc://{username}:{password}@{server}/{database}?driver={driver}"
        engine = create_engine(connection_string)
        return(engine)
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)
# Función para limpiar caracteres especiales en los nombres de las columnas
def LimpiarCaracteres(Tabla):
    try:
        Tabla.columns = Tabla.columns.str.replace('á','a')
        Tabla.columns = Tabla.columns.str.replace('Á','a')
        Tabla.columns = Tabla.columns.str.replace('é','e')
        Tabla.columns = Tabla.columns.str.replace('É','E')
        Tabla.columns = Tabla.columns.str.replace('í','i')
        Tabla.columns = Tabla.columns.str.replace('Í','I')
        Tabla.columns = Tabla.columns.str.replace('ó','o')
        Tabla.columns = Tabla.columns.str.replace('Ó','O')
        Tabla.columns = Tabla.columns.str.replace('ú','u')
        Tabla.columns = Tabla.columns.str.replace('Ú','U')
        #Se aplica la limpieza a todas las filas de la tabla 
        Tabla = Tabla.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        Tabla = Tabla.replace('', np.nan)
        return Tabla
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)
    
def Limpieza(Diccionario):
    try:
        rutaOne=r'{}'.format(Diccionario.get('Anexo5OneDrive'))
        One=pd.read_excel(rutaOne)
        One.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        wb = load_workbook(rutaOne)
        page = wb['Negociantes']
        columnas_a_limpiar = ['F', 'G', 'H', 'I']
        # Iterar sobre las filas de las columnas seleccionadas, comenzando desde la segunda fila
        for columna in columnas_a_limpiar:
            for celda in page[columna][1:]:
                celda.value = None
        wb.save(rutaOne)
        return 'Limpieza Realizada'
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)

# Función para cargar un insumo desde un archivo Excel
def CargueInsumo(Diccionario):
    try:
        #Lectura de los archivos excel
        engine=ConexionDB(Diccionario)
        ruta=r'{}'.format(Diccionario.get('Ruta'))
        ruta=os.path.normpath(ruta)
        hoja=str(Diccionario.get('Hoja'))
        Tabla = pd.read_excel(ruta, sheet_name=hoja)
        # Se eliminan los espacios en las tablas al inicio y final 
        Tabla = Tabla.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        # Se reemplazan valores en blanco por valores nulos 
        if(hoja!='Pedidos'):
            Tabla = Tabla.drop_duplicates()
        Tabla = Tabla.replace('', np.nan)
        if hoja == 'Negociantes':
            #Filtros a la tabla Negociantes
            Tabla = Tabla[(Tabla['Procesar'] == 'SI')|(Tabla['Procesar'] == 'VACACIONES')]
            Negociantes = Tabla[(Tabla['Jefe de Unidad'] !='ANALISTA DE COMPRAS')&(Tabla['Jefe de Unidad'] !='INSTITUCIONAL')&(Tabla['Negociador']!='ANALISTA DE COMPRAS')]
            Negociantes = Negociantes[['Negociador','Correo Negociador', 'Jefe de Unidad', 'Correo Jefe de Unidad']]
            Nulos = Negociantes.isna().sum().sum()     
        else:
            Nulos = Tabla.isna().sum().sum()
        if (Nulos == 0 or hoja=='Pedidos'):
            if 'VALOR' in Tabla.columns:
                Tabla['VALOR'] = Tabla['VALOR'].round().astype(float)
            Tabla.to_sql(hoja, con=engine, if_exists='replace', schema='NotasCreditoYFacturacion', index=False)
            return ('Cargue Exitoso')
        else:
            return ('Tabla con valores nulos')
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)     


#def CargueIndicadores(Diccionario):

# Función para cargar un correo desde un archivo Excel
def CargueCorreo(Diccionario):
    try:
        ruta=r'{}'.format(Diccionario.get('rutaAdjunto'))
        ruta=os.path.normpath(ruta)
        hoja=str(Diccionario.get('Hoja'))
        Tabla = pd.read_excel(ruta, sheet_name=hoja)
        # Se reemplazan valores en blanco por valores nulos 
        engine=ConexionDB(Diccionario)
        Tabla=LimpiarCaracteres(Tabla)
        # Validación de la estructura de la tabla
        Validacion=Tabla.columns[0]=='NIT' and Tabla.columns[2]=='VALOR' and Tabla.columns[3]=='CONCEPTO' and Tabla.columns[4]=='DETALLE' and Tabla.columns[5]=='CATEGORIA' and Tabla.columns[6]=='NEGOCIADOR' and Tabla.columns[7]=='UN. NEGOCIO' and Tabla.columns[9]=='PROCESAR'
        if Validacion == False:
            Validacion=Tabla.columns[0]=='NIT' and Tabla.columns[3]=='VALOR' and Tabla.columns[4]=='CONCEPTO' and Tabla.columns[5]=='DETALLE' and Tabla.columns[6]=='CATEGORIA' and Tabla.columns[7]=='NEGOCIADOR' and Tabla.columns[8]=='UN. NEGOCIO' and Tabla.columns[10]=='PROCESAR'

        if (Validacion):
            Tabla = Tabla[['NIT','VALOR','CONCEPTO','DETALLE','CATEGORIA','NEGOCIADOR','UN. NEGOCIO','TIPO NOTA/FACTURA','PROCESAR']]
            # Verificación de que no hay valores en cero y nulos
            Ceros=((Tabla['VALOR']==0).sum())==0
            Nulos=Tabla.isna().sum().sum()==0
            ValNum = pd.to_numeric(Tabla['VALOR'], errors='coerce').isna().sum()==0&pd.to_numeric(Tabla['NIT'], errors='coerce').isna().sum()==0
            
            if (Nulos and Ceros and ValNum):
                Tabla['VALOR']=Tabla['VALOR'].round().astype(float)
                Tabla.to_sql('PreConsolidado', con=engine, if_exists='append', schema='NotasCreditoYFacturacion', index=False)
                return ('Cargue Exitoso')
            else:
                # Mensajes de error según las validaciones
                if not Nulos:
                    return ("El archivo enviado contiene valores nulos")
                elif not ValNum:
                    return ("Existen valores de texto en las columnas NIT o VALOR")
                else:
                    return ("Existen valores en 0")
        else:
            return ("La estructura no es correcta")
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        error= 'El archivo no pudo ser leido' + str(e)
        return  error

#Funcion para crear archivos segun plantilla
def CrearArchivos(Diccionario):
    try:
        engine=ConexionDB(Diccionario)
        Query=Diccionario.get("Query")
        filename=r'{}'.format(Diccionario.get('filename'))
        Tabla=pd.read_sql_query(sql=text(Query), con=engine.connect())
        Tabla['VALOR'] = Tabla['VALOR'].astype(float)
        Tabla['VALOR'] = Tabla['VALOR'].astype(int)    
        #Tabla['VALOR'] = Tabla['VALOR'].round().astype(int)
        wb = load_workbook(filename)
        page = wb['CIERRE']
        start_row = 2  # Fila de inicio (A2)
        start_col = 1  # Columna de inicio (A)
        for r in dataframe_to_rows(Tabla, index=False, header=False):
            for idx, cell_value in enumerate(r):
                page.cell(row=start_row, column=start_col+idx, value=cell_value)
            start_row += 1
        wb.save(filename)
        return ("Archivo Actualizado")
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)


#Funcion para actualizar Anexo 5 y base de datos en la tabla Negociantes
def Actualizar(Diccionario):
    try:
        engine=ConexionDB(Diccionario)
        rutaOne=r'{}'.format(Diccionario.get('Anexo5OneDrive'))
        rutaFile=r'{}'.format(Diccionario.get('Anexo5'))
        rutaFile=os.path.normpath(rutaFile)
        File=pd.read_excel(rutaFile)
        One=pd.read_excel(rutaOne)
        One.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        File.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        File=File[['Negociador','Correo Negociador','Jefe de Unidad','Correo Jefe de Unidad','Procesar','Estado Recibido','Asunto Correo','Estado Aprobacion']]
        One=One[['Negociador','Id Aprobación', 'Estado Aprobacion']]
        Negociantes=pd.merge(File,One,left_on='Negociador',right_on='Negociador')
        Negociantes.to_sql('Negociantes',con=engine, schema='NotasCreditoYFacturacion', if_exists='replace')
        wb = load_workbook(rutaOne)
        page = wb['Negociantes']
        fila=2
        for r in dataframe_to_rows(Negociantes, index=False, header=False):
            c2=page.cell(row=fila,column=1,value=r[0])
            c2=page.cell(row=fila,column=2,value=r[1])
            c2=page.cell(row=fila,column=3,value=r[2])
            c2=page.cell(row=fila,column=4,value=r[3])
            c2=page.cell(row=fila,column=5,value=r[4])
            c2=page.cell(row=fila,column=6,value=r[5])
            c2=page.cell(row=fila,column=7,value=r[6])
            c2=page.cell(row=fila,column=8,value=r[8])
            if ((r[7]=='Pendiente') and (r[9]=='Rechazado')):
                c2=page.cell(row=fila,column=9,value=r[7])
            else:
                c2=page.cell(row=fila,column=9,value=r[9])
            fila=fila+1
        wb.save(rutaFile)
        return 'Tabla Actualizada'
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)
#Funcion para eliminar duplicados 
def DropDuplicates(Diccionario):
    try:
        engine=ConexionDB(Diccionario)
        Query=Diccionario.get("Query")
        Tb=Diccionario.get("Tabla")
        Tabla=pd.read_sql_query(sql=text(Query), con=engine.connect())
        Tabla['VALOR']=Tabla['VALOR'].astype(int)
        Tabla=Tabla.drop_duplicates()
        Tabla.to_sql(Tb,con=engine,if_exists='replace',schema='NotasCreditoYFacturacion',index=False)
        return 'Duplicados eliminados'
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)

def Export (Diccionario):
    try:
        # Leer el archivo Excel
        engine=ConexionDB(Diccionario)
        Query=Diccionario.get("Query")
        filename=r'{}'.format(Diccionario.get('filename'))
        Tabla=pd.read_sql_query(sql=text(Query),con=engine.connect())
        Tabla.iloc[:,10]='0206'
        Tabla['valor']=Tabla['valor'].astype('int')
        Tabla['NIT']=Tabla['NIT'].astype('int')
        # Definir los anchos de cada columna
        TamañoCol = [10,4,4,2,2,18,10,4,8,1,4,40,160,4,14]
        # Crear una cadena de texto formateada con los valores de cada fila
        FormatFilas = []
        for _, row in Tabla.iterrows():
            FormatoFila = ''
            for col, width in zip(row, TamañoCol):
                FormatoCol = str(col)[:width].ljust(width)
                FormatoFila += FormatoCol
            FormatFilas.append(FormatoFila)
        # Unir las filas formateadas en una sola cadena de texto con saltos de línea
        Salida = '\n'.join(FormatFilas)
        # Guardar la cadena de texto en un archivo
        with open(filename, 'w') as f:
            f.write(Salida)
        return 'Creacion Exitosa'
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)

def NotaCredito(Diccionario):
    try:
        engine = ConexionDB(Diccionario)
        Query='SELECT * FROM NotasCreditoYFacturacion.NOTAS ORDER BY ISNULL(NoNOTA, 99999) ASC;'
        #Query2='SELECT * FROM NotasCreditoYFacturacion.NOTAS ORDER BY ISNULL(NoNOTA, 99999) ASC;'
        ruta = r'{}'.format(Diccionario.get('Ruta'))
        fecha=Diccionario.get('Fecha')
        fechaYM=Diccionario.get('FechaYM')
        rutaFinal=Diccionario.get('RutaResultados')
        U_Primera_Ejecución=Diccionario.get('DiasPrimeraEjecución')
        if(datetime.now().day<30):
            ultimoDiaMesAnterior=datetime.now().replace(day=1) - timedelta(days=1)
            #ultimoDiaMesAnterior=datetime.date.today().replace(day=1)+datetime.timedelta(days=-1)
        else:
            ultimoDiaMesAnterior= datetime.now()
        print(ultimoDiaMesAnterior)
        fechaColum=int(fechaYM + str(ultimoDiaMesAnterior.day)) 
        Tabla=pd.read_sql_query(sql=text(Query),con=engine.connect())
        #Tabla['NIT']=Tabla['NIT'].astype('int')
        for i, r in Tabla.iterrows():
            if pd.isnull(r['NoNOTA']):
                if i==0:
                    Tabla.at[i,'NoNOTA']='01-'+fecha+'-001'
                else:
                    numNota=Tabla.iloc[i-1]['NoNOTA']
                    lsConsec=numNota.split('-')
                    if int(lsConsec[3])<999:
                        lsConsec[0]=str(lsConsec[0]).zfill(2)
                        lsConsec[3]=str(int(lsConsec[3])+1).zfill(3)
                        Tabla.at[i,'NoNOTA']=lsConsec[0]+'-'+lsConsec[1]+'-'+lsConsec[2]+'-'+lsConsec[3] 
                    elif int(lsConsec[3])==999:
                        lsConsec[0]=str(int(lsConsec[0])+1).zfill(2)
                        lsConsec[3]='001'
                        Tabla.at[i,'NoNOTA']=lsConsec[0]+'-'+lsConsec[1]+'-'+lsConsec[2]+'-'+lsConsec[3]
        Tabla = Tabla[Tabla['PROCESAR']=='SI']           
        Tabla.to_sql('NotasCons',con=engine,if_exists='replace',schema='NotasCreditoYFacturacion',index=False)
        #Tabla.to_excel(rutaFinal+'NotasCredito'+fechaYM+'.xlsx', index=False)
        Tabla=Tabla[['NIT','NoNOTA','CONCEPTO','VALOR','CLASE','JERARQUIA','CUENTA']]
        # Separacion de archivos de tipo xlsx de grupos de 500 datos 
        cantidad = 500
        num_registros = len(Tabla)
        num_archivos = math.ceil(num_registros / cantidad )
        for i in range(num_archivos):
            inicio = i * cantidad
            fin = (i + 1) * cantidad
            registros = Tabla.iloc[inicio:fin]
            num_datos= len(registros)
            registros.loc[:, 'NIT'] = registros['NIT'].astype(float).apply(lambda x: str(int(x)) if x.is_integer() else str(x))
            #registros.loc[:, 'NIT'] = registros['NIT'].astype(int).astype(str).str.rstrip('.0')
            total = registros['VALOR'].sum()
            registros.insert(0,'colum0',2205)
            registros.insert(4,'colum4',1000000)
            registros.insert(8,'colum8',0)
            registros.insert(9,'colum9','COMS')
            registros.insert(10,'colum10',fechaColum-i)
            registros.insert(11,'colum11',fechaColum-i)
            #registros['DETALLE']=registros['DETALLE'].replace({'\n':''},regex=True)
            #registros['DETALLE']=registros['DETALLE'].replace({'\n':''},regex=True)
            #registros['DETALLE']=registros['DETALLE'].replace({'\n':''},regex=True)
            #registros['DETALLE']=registros['DETALLE'].replace({',':''},regex=True)
            #registros['DETALLE']=registros['DETALLE'].replace({'á':'a'},regex=True)
            fila0=pd.DataFrame([{}])
            registros= pd.concat([fila0,registros]).reset_index(drop=True)
            registros.iloc[0, 0] = 0
            registros.iloc[0, 1] = 'EGRS'
            registros.iloc[0, 2] = fechaColum-i
            registros.iloc[0, 3] = num_datos
            registros.iloc[0, 4] = total
            #registros.to_excel(archivo_destino, index=False, header=False)
            #archivo_destino = ruta + f'Notas Credito ' + str(i+1) + '.csv'
            archivo_destino = ruta + str(registros.iloc[0, 2]) + '.csv'
            #print(archivo_destino)
            #registros.to_excel(archivo_destino, index=False, header=False)
            registros=LimpiarCaracteres(registros)
            registros.to_csv(archivo_destino, index=False, header=False, sep=';', float_format='%.0f', encoding='latin-1')
            registros=pd.read_csv(archivo_destino, sep='|', header=None, encoding='latin-1')
            registros.loc[0,0]=registros.loc[0,0].rstrip(';')
            registros.to_csv(archivo_destino, index=False, header=False, sep='|', float_format='%.0f', encoding='latin-1')
                #total2.to_excel(rutaFinal+'NotasCredito'+fechaYM+'.xlsx', index=False)
        
        return ("Archivos separados "+str(num_archivos))
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)    

def GenerarListaPedidos(Diccionario):
    try:
        rutaCSV=r'{}'.format(Diccionario.get('rutaCSV'))
        rutaExcel=r'{}'.format(Diccionario.get('rutaExcel'))
        print(rutaCSV)
        print(rutaExcel)
        engine=ConexionDB(Diccionario)
        df=pd.read_csv(rutaCSV, encoding='latin-1',sep='\t',skiprows=4)
        df=df.reset_index()
        df=df.iloc[:, [1,5,17,22,24,33,79]]
        df.columns=['Doc.comer.','Denominacion','Solic.','Material','Nombre 1','Valor neto','Factura']
        df['Valor neto']=pd.to_numeric(df['Valor neto'].str.replace('.', ''))
        df.to_excel(rutaExcel,sheet_name='Pedidos')
        df.to_sql('Pedidos',con=engine,if_exists='replace',schema='NotasCreditoYFacturacion',index=False)
        return('Exitoso')
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        line_number = exc_traceback.tb_lineno
        e = f"Error en la linea {line_number}, {e}"
        return str(e)

print(NotaCredito(Diccionario))

