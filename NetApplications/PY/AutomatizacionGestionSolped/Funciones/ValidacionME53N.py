# ============================================
# Función Local: ValidacionME53N
# Autor: Paula Sierra - NetApplications
# Descripcion: Funciones de validacion
# Ultima modificacion: 02/02/2026
# Propiedad de Colsubsidio
# Cambios:
# ============================================
import traceback
import win32com.client
import time
import os
from Funciones.EscribirLog import WriteLog
from Config.settings import RUTAS
import pandas as pd
import datetime
import re
import win32clipboard
import pyautogui
import chardet
from datetime import datetime
from typing import Dict, List, Tuple
import smtplib
import os
from Funciones.EmailSender import EmailSender
from typing import List, Union
import sys
from openpyxl import load_workbook
from Funciones.ValidacionHU3 import ValidarContraTabla


def DeterminarEstadoFinal(datosTexto: Dict, validaciones: Dict) -> Tuple[str, str]:
    """
    Determina el estado final y observaciones basado en validaciones
    AJUSTADO: Maneja textos que solo son descripciones sin datos estructurados
    """
    # Cortar validación temprana para textos sin estructura
    if datosTexto.get("tipo_texto") == "solo_descripcion":
        return "Solo descripcion", "El texto solo contiene una descripción del producto"

    if datosTexto.get("tipo_texto") == "tabla_sap":
        return "Texto invalido", "El texto contiene una tabla SAP exportada"

    if datosTexto.get("tipo_texto") == "vacio":
        return "Sin Texto", "El item no tiene texto"

    campos_obligatorios_presentes = validaciones.get("campos_obligatorios", {}).get(
        "presentes", 0
    )
    total_campos_obligatorios = validaciones.get("campos_obligatorios", {}).get(
        "total", 4
    )
    campos_validados = validaciones.get("campos_validados", 0)

    # CASO 1: Texto vacio o muy corto
    concepto = datosTexto.get("concepto_compra", "")
    if not concepto or len(concepto.strip()) < 5:
        return "Sin Texto", "No se encontro texto en el item"

    # CASO 2: Texto es solo tabla de SAP (detectar por pipes y guiones)
    if concepto.count("|") > 10 and concepto.count("-") > 50:
        return (
            "Texto invalido",
            "El texto es una tabla de SAP exportada, no contiene informacion del proveedor",
        )

    # CASO 3: Texto es solo descripcion del producto (sin datos del proveedor)
    # Si NO tiene ningun campo obligatorio Y el texto es corto (menos de 200 chars)
    if campos_obligatorios_presentes == 0 and len(concepto) < 200:
        return (
            "Solo descripcion",
            f"Texto solo contiene descripcion del producto: {concepto[:50]}...",
        )

    # CASO 4: Texto tiene algunos datos pero incompletos
    if campos_obligatorios_presentes == 0 and len(concepto) >= 200:
        return (
            "Verificar manualmente",
            "Texto extenso pero sin campos estructurados (NIT, valores, etc)",
        )

    # CASO 5: Validacion normal - tiene campos estructurados
    if campos_obligatorios_presentes >= 3 and campos_validados >= 3:
        estado = "Registro validado para orden de compra"
        observaciones = "Validacion exitosa - Cumple requisitos minimos"
    elif campos_obligatorios_presentes >= 2:
        estado = "Verificar manualmente"
        observaciones = GenerarObservaciones(datosTexto, validaciones)
        if campos_validados < 2:
            estado = "Datos no coinciden con SAP"
    else:
        estado = "Falta informacion critica"
        observaciones = GenerarObservaciones(datosTexto, validaciones)

    # =============================================================
    #  VALIDACIÓN DE CAMPOS OBLIGATORIOS SAP (ME53N)
    #  SI FALTAN → ESTADO = "Verificar manualmente"
    # =============================================================
    campos_me = validaciones.get("campos_me53n", {})

    if campos_me and campos_me.get("faltantes"):
        faltantes = campos_me["faltantes"]

        # agregar a observaciones
        if observaciones:
            observaciones += " | "

        observaciones += f"Faltan campos SAP: {', '.join(faltantes)}"

        # cambiar estado final
        return "Verificar manualmente", observaciones

    return estado, observaciones


def ExtraerDatosTexto(texto: str) -> Dict:
    """Extrae campos estructurados del texto capturado
    AJUSTADO: Se agregan nuevos patrones manteniendo la lógica existente."""

    datos = {
        "razon_social": "",
        "nit": "",
        "correo": "",
        "empresa": "",
        "concepto_compra": "",
        "fecha_prestacion": "",
        "valor_unitario": "",
        "valor_total": "",
        "cantidad": "",
        "subtotal": "",
        "iva_impo": "",
        "total": "",
        "responsable_compra": "",
        "ceco": "",
        "telefono": "",
        "direccion_entrega": "",
        "tipo_texto": "desconocido",
    }

    if not texto or not texto.strip():
        datos["tipo_texto"] = "vacio"
        return datos

    texto_limpio = texto.strip()
    texto_upper = texto_limpio.upper()
    lineas = [linea.strip() for linea in texto_limpio.split("\n") if linea.strip()]

    # ------------------------------------------------------------
    # CLASIFICACIÓN DEL TEXTO (MANTENIDA TAL CUAL)
    # ------------------------------------------------------------

    if texto.count("|") > 10 and texto.count("-") > 50:
        datos["tipo_texto"] = "tabla_sap"
        datos["concepto_compra"] = "TABLA SAP EXPORTADA"
        return datos

    if len(texto_limpio) < 200 and not any(
        kw in texto_upper
        for kw in ["NIT", "RAZON SOCIAL", "VALOR", "CANTIDAD:", "PROVEEDOR"]
    ):
        datos["tipo_texto"] = "solo_descripcion"
        datos["concepto_compra"] = texto_limpio
        return datos

    if any(
        kw in texto_upper
        for kw in ["NIT", "RAZON SOCIAL", "PROVEEDOR:", "VALOR TOTAL", "CANTIDAD:"]
    ):
        datos["tipo_texto"] = "estructurado"
    else:
        datos["tipo_texto"] = "texto_simple"

    # ------------------------------------------------------------
    # EXTRACCIÓN DE CAMPOS (NUEVOS + EXISTENTES)
    # ------------------------------------------------------------

    # --- RAZON SOCIAL (NUEVO + mantiene lo que tenías) ---
    m = re.search(r"GENERAR ORDEN DE COMPRA A[:\s]*(.+?)\s*NIT", texto_upper)
    if m:
        datos["razon_social"] = m.group(1).strip()

    # fallback TUYO (línea con RAZON SOCIAL o PROVEEDOR)
    if not datos["razon_social"]:
        for linea in lineas:
            linea_upper = linea.upper()
            if any(kw in linea_upper for kw in ["RAZON SOCIAL", "PROVEEDOR"]):
                if ":" in linea:
                    datos["razon_social"] = linea.split(":", 1)[1].strip()
                    break

    # --- NIT (MANTENIDO + más flexible) ---
    patrones_nit = [r"NIT[\s:]*([0-9.\-]+)", r"IDENTIFICACION[\s:]*([0-9.\-]+)"]
    for patron in patrones_nit:
        m = re.search(patron, texto_upper)
        if m:
            datos["nit"] = m.group(1).strip()
            break

    # --- CORREOS (MEJORA) ---
    correos = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", texto)

    # Correo del proveedor (NO colsubsidio)
    correos_proveedor = [c for c in correos if "colsubsidio.com" not in c.lower()]
    if correos_proveedor:
        datos["correo"] = correos_proveedor[0].strip()

    # Correos responsables (@colsubsidio.com)
    correos_resp = [c.lower() for c in correos if "colsubsidio.com" in c.lower()]
    if correos_resp:
        datos["responsable_compra"] = ", ".join(correos_resp)

    # --- CONCEPTO (NUEVO + tu fallback) ---
    m = re.search(
        r"POR CONCEPTO DE[:\s]*(.+?)\s*(EMPRESA|FECHA|HORA|DIRECCION|CANTIDAD|MES|HTH)",
        texto_upper,
    )
    if m:
        datos["concepto_compra"] = m.group(1).strip()

    # fallback TUYO
    if not datos["concepto_compra"] and lineas:
        datos["concepto_compra"] = lineas[0][:200]

    # --- CANTIDAD (MANTENIDO) ---
    m = re.search(r"CANTIDAD[\s:]*([0-9.,]+)", texto_upper)
    if m:
        datos["cantidad"] = m.group(1).strip()

    # --- VALORES (MANTENIDO) ---
    patron_valor = r"[\$]?\s*([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)"

    # valor unitario
    for linea in lineas:
        if any(
            kw in linea.upper()
            for kw in [
                "VALOR UNITARIO",
                "VALOR UNIDAD",
                "VR UNITARIO",
                "PRECIO UNITARIO",
            ]
        ):
            m = re.search(patron_valor, linea)
            if m:
                datos["valor_unitario"] = m.group(1).strip()
                break

    # --- IVA (MEJORADO) ---
    m = re.search(r"IVA[\s:]*" + patron_valor, texto_upper)
    if m:
        datos["iva_impo"] = m.group(1).strip()

    # valor total / subtotal (lógica según IVA)
    if datos["iva_impo"]:  # Si hay IVA, buscar valor SIN IVA
        for linea in lineas:
            if any(kw in linea.upper() for kw in ["VALOR SIN IVA", "SUBTOTAL"]):
                m = re.search(patron_valor, linea)
                if m:
                    datos["valor_total"] = m.group(1).strip()
                    break
    else:  # Si NO hay IVA, tomar VALOR TOTAL normal
        for linea in lineas:
            if any(kw in linea.upper() for kw in ["VALOR TOTAL", "SUBTOTAL"]):
                m = re.search(patron_valor, linea)
                if m:
                    datos["valor_total"] = m.group(1).strip()
                    break

    # --- CECO (NUEVO) ---
    m = re.search(r"CECO[:\s]*([0-9]+)", texto_upper)
    if m:
        datos["ceco"] = m.group(1).strip()

    return datos


def GenerarObservaciones(datosTexto: Dict, validaciones: Dict) -> str:
    """Genera observaciones detalladas
    AJUSTADO: Incluye informacion sobre el tipo de texto"""

    observaciones = []

    # Agregar info sobre tipo de texto
    tipo_texto = datosTexto.get("tipo_texto", "desconocido")
    if tipo_texto == "solo_descripcion":
        return "Texto solo contiene descripcion del producto - No incluye datos del proveedor"
    elif tipo_texto == "tabla_sap":
        return "Error: Texto es una tabla de SAP, no informacion del proveedor"
    elif tipo_texto == "vacio":
        return "Item sin texto"
    elif tipo_texto == "texto_simple":
        observaciones.append("Texto no estructurado")

    # Campos obligatorios faltantes
    if "campos_obligatorios" in validaciones:
        campos_faltantes = validaciones["campos_obligatorios"].get("faltantes", [])
        if campos_faltantes:
            observaciones.append(f"Faltan: {', '.join(campos_faltantes)}")

    # Validaciones fallidas
    campos_validacion = {
        "cantidad": "Cantidad",
        "valor_unitario": "Valor unitario",
        "valor_total": "Valor total",
    }

    for campo, nombre in campos_validacion.items():
        if campo in validaciones and isinstance(validaciones[campo], dict):
            if validaciones[campo].get("texto") and not validaciones[campo].get(
                "match", False
            ):
                diferencia = validaciones[campo].get("diferencia", "")
                if diferencia:
                    observaciones.append(f"{nombre} {diferencia}")

    if not observaciones:
        return "Texto sin campos requeridos para validacion"

    return " | ".join(observaciones[:5])


def GenerarReporteValidacion(
    solped: str,
    item: str,
    datosTexto: Dict,
    validaciones: Dict,
    tieneAttachments: bool = None,
    obsAttachments: str = "",
    archivosAdjuntosLista: list = None,
) -> str:
    """Genera un reporte legible de la validacion"""

    reporte = f"\n{'='*80}\n"
    reporte += f"REPORTE DE VALIDACION - SOLPED: {solped}, ITEM: {item}\n"
    reporte += f"{'='*80}\n\n"

    # ===================================================================================
    # VALIDACIÓN ATTACHMENT LIST DE LA SOLPED:
    # ===================================================================================
    if tieneAttachments is False:
        # NO TIENE ADJUNTOS
        reporte += (
            "ADVERTENCIA: La SOLPED NO tiene archivos adjuntos.\n"
            f"    Detalle: {obsAttachments}\n"
            "    La validación del ítem continúa, pero la SOLPED puede ser rechazada.\n\n"
        )

    elif tieneAttachments is True:
        # SÍ TIENE ADJUNTOS → mostrar lista
        reporte += (
            "INFORMACIÓN DE ATTACHMENTS:\n"
            "    La SOLPED cuenta con archivos adjuntos.\n"
        )

        # Si existe lista estructurada de archivosAdjuntos
        if archivosAdjuntosLista:
            reporte += (
                f"    Total de archivos: {len(archivosAdjuntosLista)}\n"
                f"    Ejemplos:\n"
            )
            # Listar hasta 3 para evitar reportes enormes
            for a in archivosAdjuntosLista[:3]:
                reporte += f"      - {a.get('title', '')}\n"

            if len(archivosAdjuntosLista) > 3:
                reporte += (
                    f"      ... ({len(archivosAdjuntosLista)-3} archivos adicionales)\n"
                )

        # Si solo viene texto plano (caso exportación)
        elif obsAttachments:
            reporte += f"    Detalle: {obsAttachments}\n"

        reporte += "\n"

    # ===================================================================================
    # 1. CAMPOS OBLIGATORIOS SAP ME53N
    # ===================================================================================
    if "campos_me53n" in validaciones:
        datosMe53n = validaciones["campos_me53n"]
        reporte += "CAMPOS OBLIGATORIOS SAP ME53N:\n"
        reporte += f"  Presentes: {datosMe53n['presentes']}/{datosMe53n['total']}\n"

        if datosMe53n["faltantes"]:
            reporte += "  Faltantes:\n"
            for campo in datosMe53n["faltantes"]:
                reporte += f"    - {campo}\n"
        else:
            reporte += "  ✓ Todos los campos obligatorios ME53N están presentes.\n"

        reporte += "\n"

    # ===================================================================================
    # 2. ADVERTENCIA POR TEXTO SIN ESTRUCTURA
    # ===================================================================================
    if datosTexto.get("tipo_texto") in ["solo_descripcion", "vacio", "tabla_sap"]:
        reporte += (
            "ADVERTENCIA IMPORTANTE:\n"
            "  Este ítem contiene solo descripción o no tiene estructura completa.\n"
            "  → No fue posible extraer todos los campos estructurados.\n"
            "  → Validar manualmente la información.\n"
            "  → Revisar cantidad, valor unitario, valor total y NIT directamente en SAP.\n\n"
        )

    # ===================================================================================
    # 3. DATOS EXTRAÍDOS DEL TEXTO
    # ===================================================================================
    reporte += "DATOS EXTRAIDOS DEL TEXTO:\n"
    reporte += f"  Razon Social: {datosTexto.get('razon_social') or 'No encontrado'}\n"
    reporte += f"  NIT: {datosTexto.get('nit') or 'No encontrado'}\n"
    reporte += f"  Correo: {datosTexto.get('correo') or 'No encontrado'}\n"
    reporte += (
        f"  Concepto: {datosTexto.get('concepto_compra')[:50] or 'No encontrado'}...\n"
    )
    reporte += f"  Cantidad: {datosTexto.get('cantidad') or 'No encontrado'}\n"
    reporte += (
        f"  Valor Unitario: {datosTexto.get('valor_unitario') or 'No encontrado'}\n"
    )
    reporte += f"  Valor Total: {datosTexto.get('valor_total') or 'No encontrado'}\n"
    reporte += (
        f"  Responsable: {datosTexto.get('responsable_compra') or 'No encontrado'}\n"
    )
    reporte += f"  CECO: {datosTexto.get('ceco') or 'No encontrado'}\n\n"

    # ===================================================================================
    # 4. CAMPOS OBLIGATORIOS EXTRAÍDOS DEL TEXTO
    # ===================================================================================
    if "campos_obligatorios" in validaciones:
        oblig = validaciones["campos_obligatorios"]
        reporte += "CAMPOS OBLIGATORIOS (Segun Texto Extraído):\n"
        reporte += f"  Presentes: {oblig['presentes']}/{oblig['total']}\n"

        if oblig["faltantes"]:
            reporte += "  Faltantes:\n"
            for campo in oblig["faltantes"]:
                reporte += f"    - {campo}\n"
        else:
            reporte += "  ✓ Todos los campos obligatorios del texto están presentes.\n"

        reporte += "\n"

    # ===================================================================================
    # 5. VALIDACIONES DETALLADAS
    # ===================================================================================
    reporte += "VALIDACIONES:\n"

    for campo, validacion in validaciones.items():

        if campo in ["resumen", "campos_obligatorios", "campos_me53n"]:
            continue  # ya fueron procesados

        if not isinstance(validacion, dict):
            continue

        if "match" not in validacion:
            continue

        estado = "EXITO" if validacion["match"] else "ERROR"
        reporte += f"  {estado} {campo.upper()}:\n"

        if validacion.get("texto"):
            reporte += f"      Texto: {validacion['texto']}\n"

        if validacion.get("tabla"):
            reporte += f"      Tabla: {validacion['tabla']}\n"

        if validacion.get("diferencia"):
            reporte += f"      {validacion['diferencia']}\n"

    # ===================================================================================
    # 6. RESUMEN FINAL
    # ===================================================================================
    reporte += f"\n{validaciones['resumen']}\n"
    reporte += f"{'='*80}\n"

    return reporte


def ProcesarYValidarItem(
    session,
    solped: str,
    item_num: str,
    texto: str,
    df_items: pd.DataFrame,
    tieneAttachments: bool = None,
    obsAttachments: str = "",
    archivosAdjuntosLista: list = None,
) -> Tuple[Dict, Dict, str, str, str]:
    """
    Procesa un item: extrae datos, valida y genera reporte
    Returns: (datosTexto, validaciones, reporte, estadoFinal, observaciones)
    """

    # 1. Extraer datos del texto
    datosTexto = ExtraerDatosTexto(texto)

    # ======================================================
    # FALLBACK: Texto SAP sin estructura (tabla_sap)
    # Usar valores desde ME53N
    # ======================================================
    if datosTexto.get("tipo_texto") == "tabla_sap":

        try:
            # Buscar la fila del item en ME53N
            fila_item = df_items[
                df_items["Pos."].astype(str).str.strip() == str(item_num).strip()
            ]

            if not fila_item.empty:
                fila_item = fila_item.iloc[0]

                datosTexto["cantidad"] = datosTexto.get("cantidad") or fila_item.get(
                    "Cantidad", ""
                )

                datosTexto["valor_unitario"] = datosTexto.get(
                    "valor_unitario"
                ) or fila_item.get("PrecioVal.", "")

                datosTexto["valor_total"] = datosTexto.get(
                    "valor_total"
                ) or fila_item.get("Valor tot.", "")

                datosTexto["concepto_compra"] = datosTexto.get(
                    "concepto_compra"
                ) or fila_item.get("Texto breve", "")

                datosTexto["observacion_texto"] = (
                    "Texto del editor SAP sin estructura. "
                    "Valores tomados directamente desde ME53N."
                )

        except Exception as e:
            print(f"⚠️ Error aplicando fallback ME53N: {e}")

    # 2. Validar contra tabla (pasando el numero de item para busqueda especifica)
    validaciones = ValidarContraTabla(datosTexto, df_items, item_num)

    # 3. Determinar estado final y observaciones
    estadoFinal, observaciones = DeterminarEstadoFinal(datosTexto, validaciones)
    # Evitar generar reportes completos cuando el texto no tiene estructura
    if datosTexto.get("tipo_texto") in ["vacio", "solo_descripcion", "tabla_sap"]:
        observaciones = (
            f"Texto sin estructura completa ({datosTexto.get('tipo_texto')}). "
            "Solo contiene descripción."
        )

    # 4. Generar reporte
    reporte = GenerarReporteValidacion(
        solped,
        item_num,
        datosTexto,
        validaciones,
        tieneAttachments,
        obsAttachments,
        archivosAdjuntosLista,
    )

    return datosTexto, validaciones, reporte, estadoFinal, observaciones


def extraerDatosReporte(fila, df, mapeo):
    """
    Extrae datos de una fila ALV SAP usando mapeo dinámico
    """
    datos = {}

    for campo_reporte, variantes in mapeo.items():
        valor = ""
        for nombre_col in variantes:
            if nombre_col in df.columns:
                valor = str(fila.get(nombre_col, "")).strip()
                if valor:
                    break
        datos[campo_reporte] = valor

    return datos


def AppendHipervinculoObservaciones(rutaExcel, carpetaReportes):
    """
    Recorre todo el Excel y agrega el hipervínculo del reporte correspondiente por SOLPED e ITEM.
    """

    wb = load_workbook(rutaExcel)
    ws = wb.active

    encabezados = [c.value for c in ws[1]]

    col_solped = encabezados.index("PurchReq") + 1
    col_item = encabezados.index("Item") + 1
    col_obs = encabezados.index("Observaciones") + 1

    for fila in range(2, ws.max_row + 1):

        solped = str(ws.cell(row=fila, column=col_solped).value).strip()
        item = str(ws.cell(row=fila, column=col_item).value).strip()

        if not solped or not item:
            continue

        ruta_reporte = os.path.join(carpetaReportes, f"Reporte_{solped}_{item}.txt")

        if not os.path.exists(ruta_reporte):
            continue

        celda_obs = ws.cell(row=fila, column=col_obs)

        texto_link = f"📄 Reporte Item {item}"

        if celda_obs.value:
            if texto_link in str(celda_obs.value):
                continue
            celda_obs.value = f"{celda_obs.value} | {texto_link}"
        else:
            celda_obs.value = texto_link

        celda_obs.hyperlink = ruta_reporte
        celda_obs.style = "Hyperlink"

    wb.save(rutaExcel)


def obtenerFilaExpSolped(dfSolpeds, solped, numeroItem):
    """
    Obtiene la fila correspondiente de dfSolpeds (expSolped03.txt)
    para un item específico

    Args:
        dfSolpeds: DataFrame con datos de expSolped03.txt
        solped: Número de SOLPED
        numeroItem: Número del item

    Returns:
        Dict con los datos de expSolped03.txt para ese item
    """
    try:
        # Buscar la fila que corresponde a esta SOLPED y este item
        mascara = (
            dfSolpeds["PurchReq"].astype(str).str.replace(".", "") == str(solped)
        ) & (dfSolpeds["Item"].astype(str).str.strip() == str(numeroItem).strip())

        filasEncontradas = dfSolpeds[mascara]

        if not filasEncontradas.empty:
            return filasEncontradas.iloc[0].to_dict()
        else:
            print(
                f"⚠️ No se encontró fila en expSolped para SOLPED {solped}, Item {numeroItem}"
            )
            return {}

    except Exception as e:
        print(f"⚠️ Error buscando en expSolped: {e}")
        return {}


def LimpiarNumeroRobusto(valor):
    """
    Convierte valores tipo SAP a float de forma segura.
    Soporta:
    2.800
    218.400
    6.994
    1.615.614
    $6.99
    """

    if valor is None:
        return 0.0

    s = str(valor).strip()

    if s == "" or s.isalpha():
        return 0.0

    # Quitar moneda y espacios
    s = s.replace("COP", "").replace("USD", "").replace("$", "").replace(" ", "")

    # Si tiene puntos y NO tiene coma → formato miles SAP
    if "." in s and "," not in s:
        s = s.replace(".", "")

    # Si tiene coma decimal
    if "," in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")

    try:
        return float(s)
    except:
        return 0.0


def ObtenerValorDesdeFila(fila, posibles_nombres, default=None):
    """
    Obtiene un valor de una fila probando múltiples nombres de columna posibles

    Args:
        fila: Dict con los datos de la fila
        posibles_nombres: Lista de nombres posibles para la columna
        default: Valor por defecto si no se encuentra

    Returns:
        El valor encontrado o el default
    """
    for nombre in posibles_nombres:
        if nombre in fila:
            valor = fila[nombre]
            if valor not in [None, "", " ", "nan"]:
                return valor
    return default
